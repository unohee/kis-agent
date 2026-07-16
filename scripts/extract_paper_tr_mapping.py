# Created: 2026-07-16
# Purpose: Extract the real->paper TR_ID mapping from the official KIS API
#          workbook so that kis_agent/core/tr_mapping.py has verifiable
#          provenance instead of a hand transcription.
# Dependencies: openpyxl
# Test Status: Exercised by tests/unit/test_paper_trading.py (skipped when the
#              workbook is absent — it is too large to track in git).

"""Extract the paper-trading TR_ID mapping from the official KIS workbook.

The workbook (`한국투자증권_오픈API_전체문서_*.xlsx`, downloadable from the KIS
API portal) is the authoritative source for which APIs paper trading serves and
under which TR_ID. It is not tracked in git because of its size, so this script
exists to make `REAL_TO_PAPER_TR` reproducible rather than trusted on faith.

Usage:
    # 현재 매핑이 문서와 일치하는지 검사 (불일치 시 exit 1)
    python scripts/extract_paper_tr_mapping.py --check

    # 문서에서 추출한 매핑을 출력
    python scripts/extract_paper_tr_mapping.py

Where the workbook's summary sheet lists several TR_IDs in one cell (e.g. buy
and sell), the per-API sheets' `tr_id` field description (row 22) carries the
`[실전투자] ... [모의투자] ...` block that pairs them up; both are parsed here.
"""

import argparse
import glob
import os
import re
import sys
from typing import Dict, Optional

# 4자 이상 영숫자 TR_ID. 웹소켓 ID(H0STCNI0)처럼 숫자가 섞인 것도 잡는다.
_TR_PATTERN = re.compile(r"\b([A-Z][0-9A-Z]{6,12})\b")

WORKBOOK_GLOB = "한국투자증권_오픈API_전체문서_*.xlsx"


def find_workbook(root: str = ".") -> Optional[str]:
    """Locate the KIS API workbook, newest first."""
    matches = sorted(glob.glob(os.path.join(root, WORKBOOK_GLOB)), reverse=True)
    return matches[0] if matches else None


def _pair_from_spec_block(text: str) -> Dict[str, str]:
    """Parse a per-API `[실전투자] ... [모의투자] ...` tr_id description block.

    The two sections list the same products in the same order, so they pair up
    positionally.
    """
    real_part, _, paper_part = text.partition("[모의투자]")
    if not paper_part:
        return {}

    def _ids(section: str):
        # ':' 가 있는 줄만 = "TTTS0308U : 일본 매수 주문" 형태의 항목
        return [
            _TR_PATTERN.findall(line)[0]
            for line in section.splitlines()
            if ":" in line and _TR_PATTERN.findall(line)
        ]

    reals, papers = _ids(real_part), _ids(paper_part)
    if len(reals) != len(papers):
        return {}
    return dict(zip(reals, papers))


def extract(workbook_path: str) -> Dict[str, str]:
    """Return the real -> paper TR_ID mapping found in the workbook."""
    import openpyxl

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    mapping: Dict[str, str] = {}

    # 1) 'API 목록' 시트: 실전 TR_ID(F열) / 모의 TR_ID(G열)
    for row in list(workbook["API 목록"].iter_rows(values_only=True))[1:]:
        real, paper = row[5], row[6]
        if not real or not paper or "미지원" in str(paper):
            continue
        reals = _TR_PATTERN.findall(str(real))
        papers = _TR_PATTERN.findall(str(paper))
        if len(reals) == len(papers):
            mapping.update(zip(reals, papers))

    # 2) 개별 API 시트: 'API 목록'이 "하단 규격서 참고"로 미룬 국가별 주문 TR
    for sheet_name in workbook.sheetnames:
        for row in workbook[sheet_name].iter_rows(values_only=True):
            for cell in row:
                if cell and "[모의투자]" in str(cell):
                    mapping.update(_pair_from_spec_block(str(cell)))

    return mapping


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check",
        action="store_true",
        help="REAL_TO_PAPER_TR가 문서와 일치하는지 검사하고 불일치 시 exit 1",
    )
    parser.add_argument("--workbook", help="워크북 경로 (기본: 저장소 루트에서 탐색)")
    args = parser.parse_args()

    workbook_path = args.workbook or find_workbook()
    if not workbook_path:
        print(
            f"워크북을 찾을 수 없습니다 ({WORKBOOK_GLOB}).\n"
            f"KIS API 포털에서 '오픈API 전체문서'를 내려받아 저장소 루트에 두세요.",
            file=sys.stderr,
        )
        return 2

    extracted = extract(workbook_path)
    print(f"# 출처: {os.path.basename(workbook_path)}", file=sys.stderr)
    print(f"# 추출된 매핑: {len(extracted)}개", file=sys.stderr)

    if not args.check:
        for real in sorted(extracted):
            print(f'    "{real}": "{extracted[real]}",')
        return 0

    from kis_agent.core.tr_mapping import REAL_TO_PAPER_TR

    wrong = {
        real: (paper, REAL_TO_PAPER_TR[real])
        for real, paper in extracted.items()
        if real in REAL_TO_PAPER_TR and REAL_TO_PAPER_TR[real] != paper
    }
    missing = {r: p for r, p in extracted.items() if r not in REAL_TO_PAPER_TR}

    for real, (doc, ours) in wrong.items():
        print(f"불일치: {real} -> 문서={doc}, 코드={ours}", file=sys.stderr)
    for real, paper in missing.items():
        print(f"누락: {real} -> {paper}", file=sys.stderr)

    if wrong or missing:
        print(f"\n실패: 불일치 {len(wrong)}건, 누락 {len(missing)}건", file=sys.stderr)
        return 1

    print("일치: REAL_TO_PAPER_TR가 공식 문서와 어긋나지 않습니다", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
