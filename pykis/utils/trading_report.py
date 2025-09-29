"""
거래내역 리포트 생성 유틸리티
==============================

생성일: 2024-08-22
목적: 계좌 거래내역을 Excel 파일로 내보내는 유틸리티 함수
의존성: pandas, openpyxl, pykis.account.api
테스트 상태: 완료

주요 기능:
- 기간별 거래내역 조회
- 종목별 거래 필터링
- 체결된 거래만 추출
- Excel 파일 생성 및 포맷팅
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..account.api import AccountAPI
from ..core.client import KISClient

logger = logging.getLogger(__name__)


class TradingReportGenerator:
    """거래내역 리포트 생성기

    계좌의 거래내역을 조회하고 Excel 파일로 내보내는 기능을 제공합니다.
    """

    def __init__(self, client: KISClient, account_info: dict):
        """
        Args:
            client: KIS API 클라이언트
            account_info: 계좌 정보 {'CANO': '...', 'ACNT_PRDT_CD': '...'}
        """
        self.client = client
        self.account_info = account_info
        self.account_api = AccountAPI(client=client, account_info=account_info)

    def get_trading_history(
        self,
        start_date: str,
        end_date: str,
        ticker: Optional[str] = None,
        only_executed: bool = True,
    ) -> pd.DataFrame:
        """거래내역 조회 및 필터링

        Args:
            start_date: 시작일 (YYYYMMDD)
            end_date: 종료일 (YYYYMMDD)
            ticker: 특정 종목코드 (None이면 전체)
            only_executed: True면 체결된 거래만, False면 전체

        Returns:
            필터링된 거래내역 DataFrame
        """
        try:
            # 거래내역 조회
            df = self.account_api.inquire_daily_ccld(
                start_date=start_date, end_date=end_date, pdno=ticker or ""
            )

            if df is None or df.empty:
                logger.info(f"거래내역이 없습니다. ({start_date} ~ {end_date})")
                return pd.DataFrame()

            # 체결된 거래만 필터링
            if only_executed:
                # tot_ccld_qty가 0보다 큰 것만 (체결수량이 있는 것)
                if "tot_ccld_qty" in df.columns:
                    df = df.copy()  # 복사본 먼저 생성
                    df["tot_ccld_qty"] = pd.to_numeric(
                        df["tot_ccld_qty"], errors="coerce"
                    ).fillna(0)
                    df = df[df["tot_ccld_qty"] > 0]

            # 데이터 타입 정리
            numeric_columns = [
                "ord_qty",
                "ord_unpr",
                "tot_ccld_qty",
                "avg_prvs",
                "ccld_amt",
            ]
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

            # 날짜 형식 정리
            if "ord_dt" in df.columns:
                df["ord_dt"] = pd.to_datetime(
                    df["ord_dt"], format="%Y%m%d", errors="coerce"
                )

            # 정렬 (날짜 내림차순)
            if "ord_dt" in df.columns:
                df = df.sort_values("ord_dt", ascending=False)

            return df

        except Exception as e:
            logger.error(f"거래내역 조회 중 오류 발생: {e}")
            return pd.DataFrame()

    def export_to_excel(
        self,
        start_date: str,
        end_date: str,
        output_path: Optional[str] = None,
        tickers: Optional[List[str]] = None,
        only_executed: bool = True,
        separate_sheets: bool = False,
    ) -> str:
        """거래내역을 Excel 파일로 내보내기

        Args:
            start_date: 시작일 (YYYYMMDD)
            end_date: 종료일 (YYYYMMDD)
            output_path: 출력 파일 경로 (None이면 자동 생성)
            tickers: 조회할 종목 리스트 (None이면 전체)
            only_executed: True면 체결된 거래만
            separate_sheets: True면 종목별로 시트 분리

        Returns:
            생성된 Excel 파일 경로
        """
        try:
            # 출력 경로 설정
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = (
                    f"trading_history_{start_date}_{end_date}_{timestamp}.xlsx"
                )

            # Excel Writer 생성
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

                if tickers and separate_sheets:
                    # 종목별로 시트 분리
                    for ticker in tickers:
                        df = self.get_trading_history(
                            start_date, end_date, ticker, only_executed
                        )

                        if not df.empty:
                            # 종목명 가져오기
                            ticker_name = (
                                df["prdt_name"].iloc[0]
                                if "prdt_name" in df.columns
                                else ticker
                            )
                            sheet_name = (
                                f"{ticker}_{ticker_name[:10]}"  # 시트명 길이 제한
                            )

                            # 데이터 쓰기
                            self._write_sheet(writer, df, sheet_name)
                            logger.info(f"시트 생성: {sheet_name} ({len(df)}건)")

                else:
                    # 전체 데이터를 하나의 시트에
                    all_data = []

                    if tickers:
                        # 특정 종목들만
                        for ticker in tickers:
                            df = self.get_trading_history(
                                start_date, end_date, ticker, only_executed
                            )
                            if not df.empty:
                                all_data.append(df)
                    else:
                        # 전체 종목
                        df = self.get_trading_history(
                            start_date, end_date, None, only_executed
                        )
                        if not df.empty:
                            all_data.append(df)

                    if all_data:
                        combined_df = pd.concat(all_data, ignore_index=True)
                        self._write_sheet(writer, combined_df, "거래내역")
                        logger.info(f"전체 거래내역: {len(combined_df)}건")
                    else:
                        # 빈 데이터프레임이라도 생성
                        self._write_sheet(writer, pd.DataFrame(), "거래내역")
                        logger.info("거래내역이 없습니다.")

                # 요약 시트 추가
                self._add_summary_sheet(writer, start_date, end_date)

            logger.info(f"Excel 파일 생성 완료: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Excel 파일 생성 중 오류 발생: {e}")
            raise

    def _write_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str):
        """Excel 시트에 데이터 쓰기 및 포맷팅

        Args:
            writer: ExcelWriter 객체
            df: 쓸 데이터프레임
            sheet_name: 시트 이름
        """
        if df.empty:
            # 빈 데이터프레임 처리
            df = pd.DataFrame(
                columns=[
                    "주문일자",
                    "주문시각",
                    "종목코드",
                    "종목명",
                    "매매구분",
                    "주문수량",
                    "주문단가",
                    "체결수량",
                    "체결단가",
                    "체결금액",
                ]
            )
        else:
            # 컬럼명 한글화
            column_mapping = {
                "ord_dt": "주문일자",
                "ord_tmd": "주문시각",
                "pdno": "종목코드",
                "prdt_name": "종목명",
                "sll_buy_dvsn_cd_name": "매매구분",
                "ord_qty": "주문수량",
                "ord_unpr": "주문단가",
                "tot_ccld_qty": "체결수량",
                "avg_prvs": "체결단가",
                "ccld_amt": "체결금액",
            }

            # 필요한 컬럼만 선택
            selected_columns = [
                col for col in column_mapping.keys() if col in df.columns
            ]
            df = df[selected_columns].copy()
            df.rename(columns=column_mapping, inplace=True)

        # Excel에 쓰기
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 워크시트 가져오기
        worksheet = writer.sheets[sheet_name]

        # 스타일 적용
        self._apply_styles(worksheet, len(df))

    def _apply_styles(self, worksheet, row_count: int):
        """Excel 스타일 적용

        Args:
            worksheet: openpyxl 워크시트
            row_count: 데이터 행 수
        """
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 테두리 스타일
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 헤더 스타일 적용
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 컬럼 너비 자동 조정
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                length + 2, 50
            )

        # 숫자 포맷 적용
        for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1):
            for cell in row:
                cell.border = thin_border

                # 숫자 컬럼 포맷
                if cell.column in [6, 7, 8, 9, 10]:  # 수량, 단가, 금액 컬럼
                    cell.number_format = "#,##0"

                # 날짜 컬럼 포맷
                if cell.column == 1:  # 주문일자
                    cell.number_format = "yyyy-mm-dd"

        # 매매구분 컬럼 색상 적용
        if "매매구분" in [cell.value for cell in worksheet[1]]:
            col_idx = None
            for idx, cell in enumerate(worksheet[1], 1):
                if cell.value == "매매구분":
                    col_idx = idx
                    break

            if col_idx:
                for row in worksheet.iter_rows(min_row=2, max_row=row_count + 1):
                    cell = row[col_idx - 1]
                    if "매수" in str(cell.value):
                        cell.font = Font(color="C00000")  # 빨간색
                    elif "매도" in str(cell.value):
                        cell.font = Font(color="0070C0")  # 파란색

    def _add_summary_sheet(
        self, writer: pd.ExcelWriter, start_date: str, end_date: str
    ):
        """요약 시트 추가

        Args:
            writer: ExcelWriter 객체
            start_date: 시작일
            end_date: 종료일
        """
        # 모든 시트의 데이터 수집
        summary_data = []

        for sheet_name in writer.sheets:
            if sheet_name != "요약":
                worksheet = writer.sheets[sheet_name]
                row_count = worksheet.max_row - 1  # 헤더 제외

                summary_data.append(
                    {
                        "시트명": sheet_name,
                        "거래건수": row_count,
                        "조회기간": f"{start_date} ~ {end_date}",
                    }
                )

        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="요약", index=False)

            # 요약 시트 스타일 적용
            worksheet = writer.sheets["요약"]
            self._apply_styles(worksheet, len(summary_df))


def generate_trading_report(
    client: KISClient,
    account_info: dict,
    start_date: str,
    end_date: str,
    output_path: Optional[str] = None,
    tickers: Optional[List[str]] = None,
    only_executed: bool = True,
    separate_sheets: bool = False,
) -> str:
    """거래내역 Excel 리포트 생성 (간편 함수)

    계좌의 거래내역을 조회하여 Excel 파일로 생성합니다.

    Args:
        client: KIS API 클라이언트
        account_info: 계좌 정보
        start_date: 시작일 (YYYYMMDD)
        end_date: 종료일 (YYYYMMDD)
        output_path: 출력 파일 경로 (None이면 자동 생성)
        tickers: 조회할 종목 리스트 (None이면 전체)
        only_executed: True면 체결된 거래만
        separate_sheets: True면 종목별로 시트 분리

    Returns:
        생성된 Excel 파일 경로

    Example:
        >>> from pykis.utils.trading_report import generate_trading_report
        >>>
        >>> # 전체 거래내역
        >>> file_path = generate_trading_report(
        ...     client, account_info,
        ...     "20250101", "20250131"
        ... )
        >>>
        >>> # 특정 종목들만, 시트 분리
        >>> file_path = generate_trading_report(
        ...     client, account_info,
        ...     "20250101", "20250131",
        ...     tickers=["005930", "000660"],
        ...     separate_sheets=True
        ... )
    """
    generator = TradingReportGenerator(client, account_info)
    return generator.export_to_excel(
        start_date=start_date,
        end_date=end_date,
        output_path=output_path,
        tickers=tickers,
        only_executed=only_executed,
        separate_sheets=separate_sheets,
    )
